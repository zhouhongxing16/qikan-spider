# -*- coding:utf-8 -*-
# -加载将会使用到的函数库
import json
import time
from urllib import parse

import requests  # 读取网页
import urllib3
import math
import re
from urllib.parse import urlencode
from lxml import etree  # 用于解析网页
from openpyxl import Workbook  # 创建表格并用于数据写入
from bs4 import BeautifulSoup  # 解析网页
import random  # 随机选择代理Ip

AllData = {}
ClassifyName = []

# --获得代理IP列表
def get_ip_list(urlip, headers2):
    web_data = requests.get(urlip, headers=headers2)
    soup = BeautifulSoup(web_data.text, 'lxml')
    ips = soup.find_all('tr')
    ip_list = []

    for k in range(1, len(ips)):
        ip_info = ips[k]
        tds = ip_info.find_all('td')
        ip_list.append(tds[1].text + ':' + tds[2].text)

    return ip_list


# -从代理IP列表里面随机选择一个
def get_random_ip(ip_list):
    proxy_list = []
    for ip in ip_list:
        proxy_list.append('http://' + ip)
    proxy_ip = random.choice(proxy_list)
    proxies = {'http': proxy_ip}
    return proxies


# -定义获取文章列表对应的链接的函数
def get_data(urllist, headers1, proxies):
    url = 'http://navi.cnki.net/knavi/Common/Search/Journal'
    index = 0
    for urli in urllist:
        print("正在爬取分类 ====> 【" + ClassifyName[index] + "】")
        data ={"StateID":"","Platfrom":"","QueryTime":"","Account":"knavi","ClientToken":"","Language":"","CNode":{"PCode":"CJFD","SMode":"","OperateT":""},"QNode":{"SelectT":"","Select_Fields":"","S_DBCodes":"","QGroup":[{"Key":"Navi","Logic":1,"Items":[],"ChildItems":[{"Key":"Journal","Logic":1,"Items":[{"Key":1,"Title":"","Logic":1,"Name":"168专题代码","Operate":"","Value":"B000?","ExtendType":0,"ExtendValue":"","Value2":""}],"ChildItems":[]}]}],"OrderBy":"OTA|DESC","GroupBy":"","Additon":""}}
        data["QNode"]["QGroup"][0]["ChildItems"][0]["Items"][0]["Value"] = urli.replace("'", "") + '?'
        data1 = {"SearchStateJson":"","displaymode":1,"pageindex":1,"pagecount":21,"index":1,"random":random.random()}
        data1["SearchStateJson"] = json.dumps(data).encode('utf-8')
        par = parse.urlencode(data1)
        print(urli+"  => " +par)
        try:
            http = urllib3.PoolManager()
            r = http.request(
                "POST",
                url,
                body=par,
                headers=headers1
            )
            #print(r.status)
            if r.status == 200:
                reponse = r.data.decode('utf-8')
                soup = BeautifulSoup(reponse, 'lxml')
                # 获取最大页数
                MaxCount = soup.find(class_='lblCount').get_text()
                pageCount = math.ceil(int(MaxCount)/21)  # 目标网页一页展示 21条数据
                print("该分类共计 "+str(MaxCount)+" 条信息" + " " + str(pageCount)+"页")
                GetOneClassInfo(url, headers1, par, proxies, pageCount)
            #print('爬虫' + str(15 * (pageend - pagestart + 1)) + '篇文章信息的第' + str(num) + '篇爬取成功！！')
        except :
            print('爬虫第' + str(i) + '页中的第' + '篇爬虫失败')
        # ---创建表格，待接收数据信息---#
        time.sleep(4)
        index += 1


def GetOneClassInfo(url, headers1, par, proxies, pageCount):
    GetOneInfo(url, headers1, par, proxies, pageCount, 1)


def GetOneInfo(url, headers1, par, proxies, pageCount, index):
    time.sleep(4)
    print("    正在处理=> 第 " + str(index) + " 页")
    #http = urllib3.ProxyManager(proxies['http'])  # 代理模式
    #http = urllib3.PoolManager()
    # 处理参数
    par = re.sub(r'pageindex=.*?&', r'pageindex='+str(index)+'&', par)

    http = urllib3.PoolManager()
    r = http.request(
        "POST",
        url,
        body=par,
        headers=headers1
    )
    if r.status == 200:
        reponse = r.data.decode('utf-8')
        soup = BeautifulSoup(reponse, 'lxml')
        detials = soup.find_all(class_='detials')

        for detail in detials:
            title = detail.find('h1').get_text()
            allp = detail.find_all('p')
            oneData = {}
            oneData["title"] = title.replace('\n', '').replace('\r', '').replace(' ', '')
            for p in allp:
                pstr = p.get_text().replace('\n', '').replace('\r', '').replace(' ', '')
                if re.match(r'[\u590d]', pstr):
                    oneData["FH"] = pstr
                elif re.match(r'[\u7efc]', pstr):
                    oneData["ZH"] = pstr
                elif re.match(r'[\u0049\u0053\u0053\u004e]', pstr):
                    oneData["ISSN"] = pstr
                elif re.match(r'[\u0043\u004e]', pstr):
                    oneData["CN"] = pstr
            #print(oneData)
    #循环遍历本子分类数据
    if index < pageCount:
        index += 1
        GetOneInfo(url, headers1, par, proxies, pageCount, index)


# wb = Workbook()  # 在内存中创建一个workbook对象，而且会至少创建一个 worksheet
# ws = wb.active  # 获取当前活跃的worksheet,默认就是第一个worksheet
# ws.cell(row=1, column=1).value = "No"
# ws.cell(row=1, column=2).value = "Title"
# ws.cell(row=1, column=3).value = "Author"
# ws.cell(row=1, column=4).value = "Institute"
# ws.cell(row=1, column=5).value = "Journal"
# ws.cell(row=1, column=6).value = "Cites"
# ws.cell(row=1, column=7).value = "Download"
# ws.cell(row=1, column=8).value = "Keywords"
# ws.cell(row=1, column=9).value = "Abstart"

# ---------------参数设置
#知网期刊爬虫（被反爬虫了）
if __name__ == '__main__':

    pagestart = 1  # 起始页
    pageend = 6  # 结束页
    keywords = '吸烟'  ### 查询的主题
    # url = 'http://search.cnki.net/search.aspx?q=' + str(keywords) + '&rank=citeNumber&cluster=all&val=CJFDTOTAL&p='
    url = 'http://navi.cnki.net/knavi/Common/LeftNavi/Journal'
    urlip = 'http://www.xicidaili.com/nt/'  # 提供代理IP的网站
    headers = {
        'Referer': 'http://navi.cnki.net/knavi/Journal.html',
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/61.0.3163.100 Safari/537.36',
        'Cookie': '15fa39ba5905f1; SID_search=201087; ASP.NET_SessionId=glrrdk550e5gw0fsyobrsr45; CNZZDATA2643871=cnzz_eid%3D610954823-1510276064-null%26ntime%3D1510290496; CNZZDATA3636877=cnzz_eid%3D353975078-1510275934-null%26ntime%3D1510290549; SID_sug=111055; LID=WEEvREcwSlJHSldRa1FhcTdWZDhML1NwVjBUZzZHeXREdU5mcG40MVM4WT0=$9A4hF_YAuvQ5obgVAqNKPCYcEjKensW4IQMovwHtwkF4VYPoHbKxJw!!',
        'Host': 'navi.cnki.net',
        'Origin': 'http://navi.cnki.net',
        'X-Requested-With': 'XMLHttpRequest'
    }

    headers2 = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.143 Safari/537.36',
    }

    headers3 = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36',
        'Cookie': 'cnkiUserKey=25580eae-42c3-a83f-73b2-b8d671760673; Ecp_ClientId=1200331200103028668; ASP.NET_SessionId=hswbijwe4b1qw2cmm1rmsist; SID_navi=1201610; Ecp_IpLoginFail=200401182.149.165.25',
        'DNT':1,
        'Host': 'navi.cnki.net',
        'Origin': 'http://navi.cnki.net',
        'Referer': 'http://navi.cnki.net/knavi/Journal.html',
        'X-Requested-With': 'XMLHttpRequest',
        'Accept':'text/plain, */*; q=0.01',
        'Accept-Encoding':'gzip, deflate',
        'Connection':'keep-alive',
        'Content-Length':1012,
        'Content-Type': 'application/x-www-form-urlencoded'
    }

    headers4 ={
    "Content-Type": "application/x-www-form-urlencoded",
    "charset":"UTF-8",
    "Connection":"close",
    "Content-Length": 1013
    }

    # ------------------------------------------
    i = 0
    urlList = []
    try:
        ## 得到一个代理 IP
        ip_list = get_ip_list(urlip, headers2)  # 获得代理IP列表
        proxies = get_random_ip(ip_list)  # 获得随机的一个代理IP


        # data1 = {"productcode": "CJFD", "random": random.random()}
        # http = urllib3.PoolManager()
        # r = http.request(
        #     "POST",
        #     url,
        #     body=parse.urlencode(data1),
        #     headers=headers
        # )
        # print(r.status)
        # if r.status == 200:
        #     reponse = r.data.decode('utf-8')
        #     print(reponse.encode('utf-8'))
        #     soup = BeautifulSoup(reponse, 'lxml')
        #     i = 0
        #     urllist = []
        #     for each_li in soup.select('li'):
        #         for each_dl in each_li.select('dl'):
        #             for each_dd in each_dl.select('dd'):
        #                 link = each_dd.find('a')
        #                 # print(link['title'] + link['onclick'].split(',')[2])
        #                 urllist.append(link['onclick'].split(',')[2])
        #                 i += 1
        #     # print("Count: " + str(i))
        #     get_data(urllist, headers4, proxies)

        # 获得每一页的文章具体文章信息页面的链接
        data = {"productcode": "CJFD", "index": 1, "random": random.random()}
        response = requests.post(url, data, headers=headers)  # 获得网页源码   ,proxies=proxies
        file = response.text.encode(response.encoding).decode('utf-8')  # 对网页信息进行转化成为可以正常现实的 UTF-8格式
        #print(file)
        soup = BeautifulSoup(file, 'lxml')

        for each_li in soup.select('li'):
            for each_dl in each_li.select('dl'):
                for each_dd in each_dl.select('dd'):
                    link = each_dd.find('a')
                    #print(link['title'] + link['onclick'].split(',')[2])
                    ClassifyName.append(link['title'])
                    urlList.append(link['onclick'].split(',')[2])
                    i += 1
        # print("Count: " + str(i))
        get_data(urlList, headers4, proxies)
    except:
        print('爬取 '+ClassifyName[i]+' 时候发生错误')

    print("====================================  最终爬取结果  ====================================")  # 打印最终结果数据
    print(json.dumps(AllData, ensure_ascii=False))  # 打印最终结果数据
    # wb.save('/Users/amao/Desktop/Smoking_CNKI.xlsx')  # 最后保存搜集到的数据
